import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import authenticate, role_guard
from app.models.attendance import Attendance
from app.models.batch import Batch
from app.models.batch_module import BatchModule
from app.models.lecture import Lecture
from app.models.marks import Marks
from app.models.module import Module
from app.models.student import Student
from app.utils.response import error_response, success_response

router = APIRouter(
    prefix="/api/reports",
    tags=["reports"],
    dependencies=[Depends(authenticate), Depends(role_guard(["ADMIN", "TRAINER"]))],
)


@router.get("/attendance")
async def attendance_report(
    batchId: str = Query(None),
    studentId: str = Query(None),
    startDate: str = Query(None),
    endDate: str = Query(None),
    current_user: dict = Depends(role_guard(["ADMIN", "TRAINER"])),
    db: Session = Depends(get_db),
):
    query = db.query(Attendance)

    if studentId:
        query = query.filter(Attendance.studentId == studentId)

    if batchId:
        lecture_ids = [l.id for l in db.query(Lecture).filter(Lecture.batchId == batchId).all()]
        if lecture_ids:
            query = query.filter(Attendance.lectureId.in_(lecture_ids))
        else:
            return success_response(data=[])

    if startDate:
        start = datetime.fromisoformat(startDate)
        query = query.filter(Attendance.markedAt >= start)

    if endDate:
        end = datetime.fromisoformat(endDate)
        query = query.filter(Attendance.markedAt <= end)

    attendances = query.all()

    data = [{
        "id": a.id,
        "studentId": a.studentId,
        "lectureId": a.lectureId,
        "status": a.status,
        "markedAt": a.markedAt.isoformat() if a.markedAt else None,
        "method": a.method,
        "student": {
            "id": a.student.id,
            "user": {
                "name": a.student.user.name,
            } if a.student.user else None,
        } if a.student else None,
        "lecture": {
            "id": a.lecture.id,
            "date": a.lecture.date.isoformat() if a.lecture.date else None,
            "module": {
                "name": a.lecture.module.name,
            } if a.lecture.module else None,
        } if a.lecture else None,
    } for a in attendances]

    return success_response(data=data)


@router.get("/marks")
async def marks_report(
    batchId: str = Query(None),
    moduleId: str = Query(None),
    current_user: dict = Depends(role_guard(["ADMIN", "TRAINER"])),
    db: Session = Depends(get_db),
):
    query = db.query(Marks)

    if moduleId:
        query = query.filter(Marks.moduleId == moduleId)

    if batchId:
        students = db.query(Student).filter(Student.batchId == batchId).all()
        student_ids = [s.id for s in students]
        if student_ids:
            query = query.filter(Marks.studentId.in_(student_ids))
        else:
            return success_response(data=[])

    marks = query.all()

    data = [{
        "id": m.id,
        "studentId": m.studentId,
        "moduleId": m.moduleId,
        "type": m.type,
        "score": m.score,
        "maxScore": m.maxScore,
        "remarks": m.remarks,
        "student": {
            "id": m.student.id,
            "user": {
                "name": m.student.user.name,
            } if m.student and m.student.user else None,
        } if m.student else None,
        "module": {
            "id": m.module.id,
            "name": m.module.name,
        } if m.module else None,
    } for m in marks]

    return success_response(data=data)


@router.get("/batch-progress")
async def batch_progress(
    batchId: str = Query(...),
    current_user: dict = Depends(role_guard(["ADMIN", "TRAINER"])),
    db: Session = Depends(get_db),
):
    batch = db.query(Batch).filter(Batch.id == batchId).first()
    if not batch:
        return error_response("Batch not found", 404)

    students = db.query(Student).filter(Student.batchId == batchId).all()
    batch_modules = db.query(BatchModule).filter(BatchModule.batchId == batchId).all()
    total_lectures = db.query(Lecture).filter(Lecture.batchId == batchId).count()

    student_progress = []
    for student in students:
        present_count = db.query(Attendance).filter(
            Attendance.studentId == student.id,
            Attendance.status == "PRESENT",
        ).count()
        attendance_pct = (present_count / total_lectures * 100) if total_lectures > 0 else 0

        avg_marks = db.query(func.avg(Marks.score)).filter(
            Marks.studentId == student.id
        ).scalar() or 0

        student_progress.append({
            "studentId": student.id,
            "studentName": student.user.name if student.user else None,
            "attendancePercentage": round(attendance_pct, 2),
            "avgMarks": round(float(avg_marks), 2),
        })

    module_progress = [{
        "moduleId": bm.moduleId,
        "moduleName": bm.module.name if bm.module else None,
        "status": bm.status,
        "completionPercent": bm.completionPercent,
    } for bm in batch_modules]

    return success_response(data={
        "batch": {
            "id": batch.id,
            "name": batch.name,
            "startDate": batch.startDate.isoformat() if batch.startDate else None,
            "endDate": batch.endDate.isoformat() if batch.endDate else None,
            "isActive": batch.isActive,
        },
        "totalStudents": len(students),
        "totalLectures": total_lectures,
        "moduleProgress": module_progress,
        "studentProgress": student_progress,
    })


@router.get("/export/{export_type}")
async def export_report(
    export_type: str,
    report: str = Query("attendance"),
    batchId: str = Query(None),
    moduleId: str = Query(None),
    studentId: str = Query(None),
    startDate: str = Query(None),
    endDate: str = Query(None),
    current_user: dict = Depends(role_guard(["ADMIN", "TRAINER"])),
    db: Session = Depends(get_db),
):
    if export_type not in ("pdf", "excel"):
        return error_response("Export type must be 'pdf' or 'excel'", 400)

    # Gather report data based on report type
    if report == "attendance":
        data = _get_attendance_data(db, batchId, studentId, startDate, endDate)
        headers = ["Student", "Date", "Status", "Method"]
    elif report == "marks":
        data = _get_marks_data(db, batchId, moduleId)
        headers = ["Student", "Module", "Type", "Score", "Max Score"]
    else:
        data = _get_attendance_data(db, batchId, studentId, startDate, endDate)
        headers = ["Student", "Date", "Status", "Method"]

    if export_type == "pdf":
        return _generate_pdf(headers, data, report)
    else:
        return _generate_excel(headers, data, report)


def _get_attendance_data(db, batch_id, student_id, start_date, end_date):
    query = db.query(Attendance)
    if student_id:
        query = query.filter(Attendance.studentId == student_id)
    if batch_id:
        lecture_ids = [l.id for l in db.query(Lecture).filter(Lecture.batchId == batch_id).all()]
        if lecture_ids:
            query = query.filter(Attendance.lectureId.in_(lecture_ids))
    if start_date:
        query = query.filter(Attendance.markedAt >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.filter(Attendance.markedAt <= datetime.fromisoformat(end_date))

    attendances = query.all()
    rows = []
    for a in attendances:
        student_name = a.student.user.name if a.student and a.student.user else "Unknown"
        date = a.lecture.date.strftime("%Y-%m-%d") if a.lecture and a.lecture.date else ""
        rows.append([student_name, date, a.status, a.method])
    return rows


def _get_marks_data(db, batch_id, module_id):
    query = db.query(Marks)
    if module_id:
        query = query.filter(Marks.moduleId == module_id)
    if batch_id:
        students = db.query(Student).filter(Student.batchId == batch_id).all()
        student_ids = [s.id for s in students]
        if student_ids:
            query = query.filter(Marks.studentId.in_(student_ids))

    marks = query.all()
    rows = []
    for m in marks:
        student_name = m.student.user.name if m.student and m.student.user else "Unknown"
        module_name = m.module.name if m.module else "Unknown"
        rows.append([student_name, module_name, m.type, str(m.score), str(m.maxScore)])
    return rows


def _generate_pdf(headers, data, report_name):
    from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"{report_name.title()} Report", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", "B", 10)
    col_width = pdf.w / (len(headers) + 1)
    for header in headers:
        pdf.cell(col_width, 8, header, border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for row in data:
        for cell in row:
            pdf.cell(col_width, 8, str(cell)[:20], border=1, align="C")
        pdf.ln()

    output = io.BytesIO()
    pdf_content = pdf.output()
    output.write(pdf_content)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={report_name}_report.pdf"},
    )


def _generate_excel(headers, data, report_name):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = report_name.title()

    ws.append(headers)
    for row in data:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_name}_report.xlsx"},
    )
